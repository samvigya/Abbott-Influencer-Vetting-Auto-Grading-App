"""
Abbott Influencer Vetting — Auto-Grader  v6.0
=============================================
Changes in v6.0:
  1. Auto-reject gate values = 10 (not 1) when triggered
  2. Multiple platform file uploads — each labelled by platform, composite grading output
  3. Reads pre-filled manual scores from existing Grading sheet in uploaded file (optional)

Core scoring rules (verified against TH_Adult_Nutrition reference):
  Auto-reject gates  : count > 0 → 10 (flag value), 0 otherwise; any = Risk Score 10
  8 Risk params      : % threshold → GREEN/AMBER/RED → 0/0.5/1.0 pts
  Relevance cols     : count > 0 → GREEN → full pts
  Pregnant           : count > 0 → RED → flag/penalty
  All columns read by NAME, never by position.
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# ─────────────────────────────────────────────────────────────────────────────
# SCORING RULES
# ─────────────────────────────────────────────────────────────────────────────

RULE_PCT      = "pct"
RULE_BINARY_R = "binary_r"   # count > 0 → RED  (auto-reject / pregnant)
RULE_BINARY_G = "binary_g"   # count > 0 → GREEN (relevance)

# (ai_column_name, grading_label, rule_type, green_max_pct, amber_max_pct, points_if_green)
COLUMN_RULES = [
    ("Substance Use Discussion",          "Substance Use",          RULE_BINARY_R, None,  None,  None),
    ("Breastfeeding",                     "Anti-breastfeeding",     RULE_BINARY_R, None,  None,  None),
    ("Vaccination",                       "Anti-vaccination",       RULE_BINARY_R, None,  None,  None),
    ("Health care stance",                "Anti-healthcare",        RULE_BINARY_R, None,  None,  None),
    ("Profanity Usage",                   "Profanity",              RULE_PCT,      0.15,  0.30,  None),
    ("Alcohol Use Discussion",            "Alcohol",                RULE_PCT,      0.05,  0.15,  None),
    ("Sensitive Visual Content",          "Sensitive content",      RULE_PCT,      0.20,  None,  None),
    ("Stereotypes or Bias",               "Stereotype & Bias",      RULE_PCT,      0.05,  0.15,  None),
    ("Violence Advocacy",                 "Violence Content",       RULE_PCT,      0.001, 0.05,  None),
    ("Political Stance",                  "Political",              RULE_PCT,      0.10,  0.25,  None),
    ("Unscientific Claims",               "Unscientific",           RULE_PCT,      0.15,  None,  None),
    ("Ultra-processed food",              "Ultra-processed Food",   RULE_PCT,      0.15,  0.30,  None),
    ("Kids Presenence",                   "Kids Presence",          RULE_BINARY_G, None,  None,  3.5),
    ("Topics around Adult Health",        "Topics Adult Health",    RULE_BINARY_G, None,  None,  5.0),
    ("Topics on Adult Healthy Nutrition", "Topics Adult Nutrition", RULE_BINARY_G, None,  None,  2.0),
    ("Media Presence & Awards",           "Awards/Media Presence",  RULE_BINARY_G, None,  None,  1.0),
    ("Brand Partnership Presence",        "Brand Partnerships",     RULE_BINARY_G, None,  None,  1.0),
    ("Pregnant",                          "Pregnant",               RULE_BINARY_R, None,  None,  None),
]

RULE_MAP         = {r[0]: r for r in COLUMN_RULES}
AUTO_REJECT_COLS = [r[0] for r in COLUMN_RULES if r[2] == RULE_BINARY_R and r[0] != "Pregnant"]
RISK_PARAM_COLS  = [r[0] for r in COLUMN_RULES if r[2] == RULE_PCT]

# ─────────────────────────────────────────────────────────────────────────────
# FORMAT DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(df):
    if "response_1" in df.columns and len(df) > 0:
        r0 = str(df["response_1"].iloc[0]).strip()
        if "HEADER" in r0.upper():   return "A"
        if "no of" in r0.lower():    return "C"
    if "Audio transcription" in df.columns and len(df) > 0:
        if "HEADER" in str(df["Audio transcription"].iloc[0]).upper(): return "IG"
    for col in ("Profanity Usage", "Kids Presenence", "Breastfeeding"):
        if col in df.columns:
            s = df[col].astype(str).str.strip().str.upper()
            if s.isin(["YES","NO"]).any(): return "B"
    if "Profanity Usage" in df.columns: return "Captions"
    return "B"

# ─────────────────────────────────────────────────────────────────────────────
# YES COUNTERS  (all by column name)
# ─────────────────────────────────────────────────────────────────────────────

def _sf(v):
    try: return float(v) if not pd.isna(v) else 0.0
    except: return 0.0

def count_a(df, col):
    if col not in df.columns: return 0.0, 0.0, 0.0
    c = _sf(df[col].iloc[0]); t = _sf(df[col].iloc[1]) if len(df)>1 else 0.0
    return c, t, (c/t if t>0 else 0.0)

def count_b(df, col):
    if col not in df.columns: return 0.0, 0.0, 0.0
    s = df[col].astype(str).str.strip().str.upper()
    c = float((s=="YES").sum()); t = float(s.isin(["YES","NO"]).sum())
    return c, t, (c/t if t>0 else 0.0)

def captions_map(df):
    m = {}
    if len(df) > 4:
        for col, val in df.iloc[4].items():
            if pd.notna(val): m[str(val).strip()] = col
    return m

def count_cap(df, col, cmap):
    pc = cmap.get(col)
    if pc is None: return 0.0, 0.0, 0.0
    c = _sf(df.iloc[0][pc]); t = _sf(df.iloc[1][pc]) if len(df)>1 else 0.0
    return c, t, (c/t if t>0 else 0.0)

def get_count(df, col, fmt, cmap=None):
    """Returns (count, total, pct) reading column by name."""
    if fmt == "Captions": return count_cap(df, col, cmap or {})
    if fmt == "B":        return count_b(df, col)
    return count_a(df, col)   # A, C, IG

# ─────────────────────────────────────────────────────────────────────────────
# CODE/POINT ROW  (pre-computed GREEN/AMBER/RED, Format A/C/IG only)
# ─────────────────────────────────────────────────────────────────────────────

def read_precomp(df, col):
    """Read pre-computed code from row 3 by column name."""
    if col not in df.columns or len(df) <= 3: return None
    v = str(df[col].iloc[3]).strip().upper()
    return v if v in ("GREEN","AMBER","RED") else None

# ─────────────────────────────────────────────────────────────────────────────
# APPLY RULE
# ─────────────────────────────────────────────────────────────────────────────

def pct_to_gar(pct, gm, am):
    if pct < gm:                          return "GREEN"
    if am is not None and pct < am:       return "AMBER"
    return "RED"

def apply_rule(count, pct, rule_type, gm, am):
    if rule_type == RULE_BINARY_R: return "RED"   if count > 0 else "GREEN"
    if rule_type == RULE_BINARY_G: return "GREEN" if count > 0 else "RED"
    return pct_to_gar(pct, gm, am if am else 999)

def score_col(df, col_name, fmt, cmap=None):
    """Score one named column. Returns (count, total, pct, final_code, mismatch_bool)."""
    rule = RULE_MAP.get(col_name)
    if rule is None: return 0.0, 0.0, 0.0, "GREEN", False
    _, _, rt, gm, am, _ = rule
    count, total, pct = get_count(df, col_name, fmt, cmap)
    computed = apply_rule(count, pct, rt, gm, am)
    if fmt in ("A","C","IG"):
        pre = read_precomp(df, col_name)
        final = computed if pre is None else pre
        mismatch = (pre is not None and pre != computed)
        return count, total, pct, final, mismatch
    return count, total, pct, computed, False

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def kids_age_summary(df, fmt):
    try:
        ac, pc = "Kids Age Group", "Kids Presenence"
        if ac not in df.columns or pc not in df.columns: return ""
        data = df.iloc[4:] if fmt in ("A","C","IG") else df
        mask = data[pc].astype(str).str.strip().str.upper() == "YES"
        ages = data.loc[mask, ac].dropna().astype(str).str.strip()
        skip = {"nan","none","none of these",""}
        return ", ".join(sorted({a for a in ages if a.lower() not in skip}))
    except: return ""

def total_vids(df, fmt, cmap=None):
    if fmt == "Captions":
        if cmap:
            col = cmap.get("Profanity Usage")
            if col and len(df)>1:
                try: return int(_sf(df.iloc[1][col]))
                except: pass
        return 0
    if fmt in ("A","C","IG"):
        for col in ("Profanity Usage","Alcohol Use Discussion","Kids Presenence","Topics around Adult Health"):
            if col in df.columns:
                try:
                    v = _sf(df[col].iloc[1])
                    if v > 0: return int(v)
                except: pass
        return 0
    for col in ("Profanity Usage","Kids Presenence","Breastfeeding"):
        if col in df.columns:
            n = df[col].astype(str).str.strip().str.upper().isin(["YES","NO"]).sum()
            if n > 0: return int(n)
    return len(df)

# ─────────────────────────────────────────────────────────────────────────────
# MANUAL SCORES READER
# Reads pre-filled manual values from the Grading sheet already in the file.
# Maps handle → {competition, topics_kids, abbott, awards, partnerships, notes}
# ─────────────────────────────────────────────────────────────────────────────

def read_manual_scores(xl):
    """
    Look for any existing Grading sheet in the file.
    Header row is row 0. Data starts at row 1.
    Handle is in the column whose pandas name is 'Username', or whose
    row-0 header is blank (first data column).
    Returns dict: handle → {competition, topics_kids, abbott, awards, partnerships, ...}
    """
    # Find best grading sheet — prefer platform-specific over master summary
    grading_df = None
    preferred_order = ["tt grading", "tiktok grading", "ig grading", "fb grading",
                       "captions grading", "grading"]
    all_grading = [(sn, df) for sn, df in xl.items() if "grading" in sn.lower()]
    if not all_grading:
        return {}

    # Try preferred order first
    for pref in preferred_order:
        for sn, df in all_grading:
            if pref in sn.lower():
                grading_df = df; break
        if grading_df is not None: break

    # Fallback: first grading sheet found
    if grading_df is None:
        grading_df = all_grading[0][1]

    # ── Find handle column ──────────────────────────────────────────────────
    # Priority 1: pandas column named 'Username' or 'IG Username'
    handle_col = None
    for col in grading_df.columns:
        if str(col).strip().lower() in ("username", "ig username", "handle"):
            handle_col = col; break

    # Priority 2: first column whose row-0 header is blank/nan (TH style)
    if handle_col is None:
        for col in grading_df.columns:
            h = str(grading_df.iloc[0][col]).strip()
            if h.lower() in ("nan","none",""):
                # Check it has handle-like values (not numbers)
                vals = grading_df[col].iloc[1:].dropna().astype(str).str.strip()
                if vals.str.match(r'^[a-z0-9_.]+$', case=False).any():
                    handle_col = col; break

    # Priority 3: second column (SG style: col1=full name, col2=handle)
    if handle_col is None and len(grading_df.columns) >= 2:
        handle_col = grading_df.columns[1]

    if handle_col is None:
        return {}

    # ── Build header map from row 0 using keyword search ──────────────────
    def find_col(keywords):
        """Find pandas column where row-0 header contains any keyword."""
        for col in grading_df.columns:
            h = str(grading_df.iloc[0][col]).lower().strip()
            if any(k in h for k in keywords):
                return col
        return None

    col_competition  = find_col(["competition", "worked with"])
    col_topics_kids  = find_col(["topics around kids"])
    col_abbott       = find_col(["abbott brand", "abbott brands"])
    col_awards       = find_col(["awards/media", "award/media", "awards /media"])
    col_partnerships = find_col(["relevant brand partner"])
    col_notes1       = find_col(["brand collaboration", "competitor"])
    col_notes2       = find_col(["relevant brand names"])
    col_award_name   = find_col(["award name"])

    def safe_read(row, col):
        if col is None: return None
        v = row[col]
        if pd.isna(v): return None
        try:
            f = float(v)
            # Return 0 as a valid value (competition=0 means "not flagged")
            return int(f) if f == int(f) else f
        except:
            s = str(v).strip()
            return s if s not in ("nan","None","") else None

    manual = {}
    for i in range(1, len(grading_df)):
        row    = grading_df.iloc[i]
        handle = str(row[handle_col]).strip()
        if not handle or handle.lower() in ("nan","none",""): continue

        entry = {
            "competition":  safe_read(row, col_competition),
            "topics_kids":  safe_read(row, col_topics_kids),
            "abbott":       safe_read(row, col_abbott),
            "awards":       safe_read(row, col_awards),
            "partnerships": safe_read(row, col_partnerships),
            "notes1":       safe_read(row, col_notes1),
            "notes2":       safe_read(row, col_notes2),
            "award_name":   safe_read(row, col_award_name),
        }
        # Store if any numeric value found (including 0) or any text value
        has_value = any(
            v is not None and (isinstance(v, (int, float)) or str(v).strip() not in ("","nan","None"))
            for v in entry.values()
        )
        if has_value:
            manual[handle] = entry

    return manual

# ─────────────────────────────────────────────────────────────────────────────
# FULL SHEET SCORER
# ─────────────────────────────────────────────────────────────────────────────

def score_sheet(df, handle, use_case):
    fmt  = detect_format(df)
    cmap = captions_map(df) if fmt == "Captions" else None
    mismatches = []

    def sc(col_name):
        count, total, pct, code, mm = score_col(df, col_name, fmt, cmap)
        if mm: mismatches.append(f"{col_name}: precomp≠computed (using count-based)")
        return count, total, pct, code

    # ── AUTO-REJECT GATES (value = 10 if flagged, 0 if clean) ────────────────
    auto_flags = {}
    for col in AUTO_REJECT_COLS:
        _, label, *_ = RULE_MAP[col]
        _, _, _, code = sc(col)
        auto_flags[label] = 10 if code == "RED" else 0   # ← 10, not 1

    auto_reject = any(v == 10 for v in auto_flags.values())

    # ── 8 RISK PARAMETERS ────────────────────────────────────────────────────
    risk_params = {}
    total_pts = 0.0
    for col in RISK_PARAM_COLS:
        _, label, _, _, _, _ = RULE_MAP[col]
        _, _, _, code = sc(col)
        pts = 1.0 if code=="RED" else (0.5 if code=="AMBER" else 0.0)
        risk_params[label] = {"code": code, "pts": pts}
        total_pts += pts

    risk_score = 10.0 if auto_reject else round((total_pts / 8) * 10, 4)

    # ── PREGNANT ──────────────────────────────────────────────────────────────
    _, _, _, preg_code = sc("Pregnant")
    preg_flag = 1 if preg_code == "RED" else 0

    # ── RELEVANCE ─────────────────────────────────────────────────────────────
    if use_case == "Pediatric":
        _, _, _, kids_code = sc("Kids Presenence")
        kids_pts  = 3.5 if kids_code == "GREEN" else 0.0
        age_sum   = kids_age_summary(df, fmt)
        auto_base = kids_pts
        relevance = {
            "Kids Presence":               {"pts": kids_pts, "code": kids_code, "auto": True},
            "Topics around kids":          {"pts": None, "manual": True, "max_pts": 3.5},
            "Abbott brands":               {"pts": None, "manual": True, "max_pts": 1.0},
            "Awards/Media Presence":       {"pts": None, "manual": True, "max_pts": 1.0},
            "Relevant Brand Partnerships": {"pts": None, "manual": True, "max_pts": 1.0},
        }
        preg_note = "Pregnant — Reject" if preg_flag else ""
    else:
        _, _, _, h_code = sc("Topics around Adult Health")
        _, _, _, n_code = sc("Topics on Adult Healthy Nutrition")
        h_pts = 5.0 if h_code=="GREEN" else 0.0
        n_pts = 2.0 if n_code=="GREEN" else 0.0
        auto_base = max(0.0, h_pts + n_pts - (2.0 if preg_flag else 0.0))
        age_sum   = ""
        relevance = {
            "Topics Adult Health":         {"pts": h_pts, "code": h_code, "auto": True},
            "Topics Adult Nutrition":      {"pts": n_pts, "code": n_code, "auto": True},
            "Abbott brands":               {"pts": None, "manual": True, "max_pts": 1.0},
            "Awards/Media Presence":       {"pts": None, "manual": True, "max_pts": 1.0},
            "Relevant Brand Partnerships": {"pts": None, "manual": True, "max_pts": 1.0},
        }
        preg_note = "Pregnant — Penalty -2pts" if preg_flag else ""

    return {
        "handle":       handle,
        "use_case":     use_case,
        "fmt":          fmt,
        "total_videos": total_vids(df, fmt, cmap),
        "auto_flags":   auto_flags,    # values are 10 or 0
        "auto_reject":  auto_reject,
        "risk_params":  risk_params,
        "risk_score":   risk_score,
        "relevance":    relevance,
        "auto_base":    auto_base,
        "preg_flag":    preg_flag,
        "preg_note":    preg_note,
        "kids_age":     age_sum,
        "mismatches":   mismatches,
    }

# ─────────────────────────────────────────────────────────────────────────────
# SHEET DETECTION & UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

PLATFORM_PREFIXES = {
    "TT ":"TT","FB ":"FB","YT ":"YT","IG + YT ":"IG_YT","IG ":"IG","Captions ":"Captions"
}
SKIP_SHEETS = {
    "master grading sheet","tt grading sheet","ig + yt grading sheet",
    "fb grading sheet","captions grading sheet","grading","profiles","unknown",
}

def detect_sheets(xl):
    grouped = {p:[] for p in PLATFORM_PREFIXES.values()}
    for sheet in xl:
        if sheet.lower().strip() in SKIP_SHEETS: continue
        for prefix, plat in PLATFORM_PREFIXES.items():
            if sheet.startswith(prefix):
                grouped[plat].append((sheet, sheet[len(prefix):].strip()))
                break
    return grouped

def detect_captions_sheets(xl_cap):
    """Captions-only file: sheets named directly by handle."""
    return [(s, s.strip()) for s in xl_cap if s.lower().strip() not in SKIP_SHEETS]

def build_profiles(xl):
    for name, df in xl.items():
        if name.lower() != "profiles": continue
        m = {}
        for _, row in df.iterrows():
            handle, full_name = None, None
            for col in df.columns:
                val = str(row[col]).strip() if pd.notna(row[col]) else ""
                if "username" in str(col).lower() and val: handle = val
                if "influencer name" in str(col).lower() and val: full_name = val
            if handle: m[handle] = full_name or handle
        return m
    return {}

def detect_use_case(xl):
    for sn, df in xl.items():
        if not any(sn.startswith(p) for p in ("TT ","FB ","IG ","IG + YT ")): continue
        if "Topics around Adult Health" not in df.columns: continue
        try:
            v = df["Topics around Adult Health"].iloc[0]
            if str(v).strip().lower() not in ("nan","none","no","") and float(v) > 0:
                return "Adult"
        except: pass
    return "Pediatric"

SKIP_HANDLE = {
    "nan","none","header","total videos processed","% prevelance","% prevalence",
    "% of occurence","code/point","verdict","no of occurences",
    "platform_username","username","",
}

def extract_handle(df):
    for col in ("platform_username","Username"):
        if col not in df.columns: continue
        for val in df[col].dropna().astype(str).str.strip():
            if val.lower() not in SKIP_HANDLE and not val.startswith("```"):
                return val
    return None

# ─────────────────────────────────────────────────────────────────────────────
# RUN SCORING FOR ONE FILE
# ─────────────────────────────────────────────────────────────────────────────

def score_file(xl, platform_label, use_case, is_captions_file=False):
    """
    Score all influencer sheets in a file.
    Returns (list_of_scores, list_of_warnings).
    platform_label: e.g. "TT", "FB", "IG", "Captions"
    """
    scores, warns = [], []

    if is_captions_file:
        sheet_list = detect_captions_sheets(xl)
    else:
        grp = detect_sheets(xl)
        sheet_list = grp.get(platform_label, [])

    for sn, h in sheet_list:
        df = xl[sn]
        try:
            s = score_sheet(df, h, use_case)
            actual = extract_handle(df)
            if actual: s["handle"] = actual
            elif is_captions_file: s["handle"] = h   # use sheet name for captions file
            scores.append(s)
        except Exception as e:
            warns.append(f"{sn}: {e}")

    return scores, warns

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLES
# ─────────────────────────────────────────────────────────────────────────────

FN = "Arial"
BG = {
    "navy":"1F3864","dkred":"C00000","brtred":"FF0000",
    "dkgrn":"375623","blue":"2E75B6","manual":"FFF2CC",
    "green":"C6EFCE","amber":"FFEB9C","red":"FFC7CE",
    "grey":"F2F2F2","name":"D9E1F2","white":"FFFFFF",
}
def fp(k):  return PatternFill("solid", fgColor=BG.get(k,k))
def bdr():
    s=Side(style="thin",color="BFBFBF"); return Border(left=s,right=s,top=s,bottom=s)
def hf(color="FFFFFF",bold=True,sz=9):   return Font(name=FN,bold=bold,color=color,size=sz)
def df_(color="000000",bold=False,sz=9): return Font(name=FN,bold=bold,color=color,size=sz)
def ac(wrap=True): return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
def al():          return Alignment(horizontal="left",  vertical="center",wrap_text=True)

# ─────────────────────────────────────────────────────────────────────────────
# GRADING SHEET WRITER
# ─────────────────────────────────────────────────────────────────────────────

def write_grading_sheet(ws, scores, label, use_case, profiles, manual_scores=None):
    """
    manual_scores: dict handle → {competition, topics_kids, abbott, awards,
                                   partnerships, notes1, notes2, award_name}
    If None or handle not found → yellow empty cells as before.
    """
    ms = manual_scores or {}

    ws.row_dimensions[1].height = 22
    t = ws.cell(1,1,f"Abbott Influencer Vetting — {label} Grading Sheet ({use_case})")
    t.font = Font(name=FN,bold=True,size=12,color=BG["navy"]); t.alignment = al()

    if use_case == "Pediatric":
        rel_cols = [
            ("Kids Presence\n(3.5 pts — AI auto)",              "dkgrn", 16, False),
            ("Topics around kids\n(3.5 pts — ⬛ Manual)",        "manual",18, True),
            ("Abbott brands\n(1 pt — ⬛ Manual)",                "manual",16, True),
            ("Awards/Media Presence\n(1 pt — ⬛ Manual)",        "manual",18, True),
            ("Relevant Brand Partnerships\n(1 pt — ⬛ Manual)",  "manual",20, True),
        ]
        notes_hdr = "Kids Age Groups (from AI)"
    else:
        rel_cols = [
            ("Topics Adult Health\n(5 pts — AI auto)",          "dkgrn", 18, False),
            ("Topics Adult Nutrition\n(2 pts — AI auto)",       "dkgrn", 18, False),
            ("Abbott brands\n(1 pt — ⬛ Manual)",                "manual",16, True),
            ("Awards/Media Presence\n(1 pt — ⬛ Manual)",        "manual",18, True),
            ("Relevant Brand Partnerships\n(1 pt — ⬛ Manual)",  "manual",20, True),
        ]
        notes_hdr = "Notes"

    COLS = [
        ("Influencer Name",                                "navy",  20,False),
        ("Handle / Username",                              "navy",  18,False),
        # Auto-reject: value = 10 or 0
        ("Substance Use\n(10 if detected — AI)",           "brtred",14,False),
        ("Worked With Competition\n(10 if yes — ⬛ Manual)","manual",20,True),
        ("Anti-breastfeeding\n(10 if detected — AI)",      "brtred",16,False),
        ("Anti-vaccination\n(10 if detected — AI)",        "brtred",15,False),
        ("Anti-healthcare\n(10 if detected — AI)",         "brtred",15,False),
        ("Auto-reject Check\n(>0 → Risk Score = 10)",      "dkred", 16,False),
        ("Profanity",                                      "blue",  12,False),
        ("Alcohol",                                        "blue",  12,False),
        ("Sensitive Content",                              "blue",  14,False),
        ("Stereotype & Bias",                              "blue",  14,False),
        ("Violence Content",                               "blue",  13,False),
        ("Political",                                      "blue",  12,False),
        ("Unscientific",                                   "blue",  13,False),
        ("Ultra-processed Food",                           "blue",  14,False),
        ("RISK Score\n(0=Clean → 10=Reject)",              "navy",  14,False),
    ] + rel_cols + [
        ("Total Relevance Score\n(max 10)",                "dkgrn", 15,False),
        ("Pregnant Or Not",                                "dkgrn", 13,False),
        ("Notes / Flags\n(⬛ Manual)",                      "manual",24,True),
        (notes_hdr,                                        "595959",28,False),
    ]

    for ci,(_,_,w,_) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ROW_H,ROW_D = 2,3
    ws.row_dimensions[ROW_H].height = 52
    for ci,(lbl,bg,_,is_m) in enumerate(COLS,1):
        c=ws.cell(ROW_H,ci,lbl); c.fill=fp(bg); c.border=bdr(); c.alignment=ac()
        c.font=hf(color="7B4F00" if is_m else "FFFFFF")

    CI_NAME=1;CI_HAND=2;CI_SUB=3;CI_COMP=4;CI_BF=5;CI_VAX=6;CI_HC=7
    CI_AR=8;CI_PROF=9;CI_ALC=10;CI_SEN=11;CI_STE=12;CI_VIO=13
    CI_POL=14;CI_UNS=15;CI_ULP=16;CI_RISK=17
    CI_R1=18;CI_R2=19;CI_R3=20;CI_R4=21;CI_R5=22
    CI_TOT=23;CI_PREG=24;CI_N1=25;CI_N2=26
    RISK_LABELS=["Profanity","Alcohol","Sensitive content","Stereotype & Bias",
                 "Violence Content","Political","Unscientific","Ultra-processed Food"]

    for off,s in enumerate(scores):
        row=ROW_D+off; rbg="white" if off%2==0 else "grey"
        ws.row_dimensions[row].height = 16
        m = ms.get(s["handle"], {})   # pre-filled manual values for this handle

        def wc(col,val,bg=None,bold=False,left=False):
            c=ws.cell(row,col,val); c.fill=fp(bg or rbg)
            c.font=df_(bold=bold); c.border=bdr()
            c.alignment=al() if left else ac(); return c

        def manual_cell(col, prefilled_val=None, hint="↑ fill manually"):
            """
            If prefilled_val is not None → show the value with light yellow bg.
            If None → empty yellow cell.
            """
            if prefilled_val is not None:
                c=ws.cell(row,col,prefilled_val)
                c.fill=fp("FFF2CC")   # still yellow so user knows it was manual
                c.font=df_(bold=True,color="375623")  # green text = pre-filled
                c.border=bdr(); c.alignment=ac()
            else:
                c=ws.cell(row,col,hint)
                c.fill=fp("manual")
                c.font=Font(name=FN,italic=True,color="7B4F00",size=8)
                c.border=bdr(); c.alignment=ac()

        # Names
        full = profiles.get(s["handle"], s["handle"])
        wc(CI_NAME,full,bg="name",bold=True,left=True)
        wc(CI_HAND,s["handle"],bg="name",left=True)

        # Auto-reject gates (AI: show 10 or 0)
        for ci,gate in [(CI_SUB,"Substance Use"),(CI_BF,"Anti-breastfeeding"),
                         (CI_VAX,"Anti-vaccination"),(CI_HC,"Anti-healthcare")]:
            v=s["auto_flags"].get(gate,0)
            wc(ci,v,bg="red" if v==10 else "green")

        # Competition — manual, read pre-filled if available
        manual_cell(CI_COMP, m.get("competition"))

        # Auto-reject sum (uses 10-based values now)
        ar_sum = sum(1 for v in s["auto_flags"].values() if v==10)
        comp_v = m.get("competition")
        if comp_v is not None and str(comp_v).strip() not in ("0","","nan"):
            ar_sum += 1   # competition flag also counts

        if ar_sum > 0:
            c=ws.cell(row,CI_AR,ar_sum)
            c.fill=fp("brtred"); c.border=bdr(); c.alignment=ac()
            c.font=Font(name=FN,bold=True,color="FFFFFF",size=9)
        else:
            wc(CI_AR,0,bg="green")

        # 8 risk params
        for ci_off,lbl in enumerate(RISK_LABELS):
            ci=CI_PROF+ci_off
            info=s["risk_params"].get(lbl,{"code":"GREEN","pts":0})
            bg="red" if info["code"]=="RED" else ("amber" if info["code"]=="AMBER" else "green")
            wc(ci,info["pts"],bg=bg)

        # Risk score (recompute if competition was pre-filled)
        rs = s["risk_score"]
        if comp_v is not None and str(comp_v).strip() not in ("0","","nan"):
            rs = 10.0   # competition flag → auto-reject
        if rs >= 5:
            c=ws.cell(row,CI_RISK,rs); c.fill=fp("brtred"); c.border=bdr(); c.alignment=ac()
            c.font=Font(name=FN,bold=True,color="FFFFFF",size=9)
        elif rs > 0:
            wc(CI_RISK,rs,bg="amber",bold=True)
        else:
            wc(CI_RISK,rs,bg="green",bold=True)

        # Relevance params
        rel_keys = list(s["relevance"].keys())
        # Map manual keys to pre-filled values
        manual_map = {
            "Topics around kids":          m.get("topics_kids"),
            "Abbott brands":               m.get("abbott"),
            "Awards/Media Presence":       m.get("awards"),
            "Relevant Brand Partnerships": m.get("partnerships"),
        }
        for i,rk in enumerate(rel_keys):
            ci=CI_R1+i
            info=s["relevance"][rk]
            if info.get("manual"):
                manual_cell(ci, manual_map.get(rk))
            else:
                pts=info["pts"] or 0
                bg="green" if pts>0 else rbg
                wc(ci,pts,bg=bg)

        # Total relevance — live SUM formula
        rl=[get_column_letter(CI_R1+i) for i in range(5)]
        pr=get_column_letter(CI_PREG)
        if use_case=="Pediatric":
            formula=f"=MIN(10,MAX(0,{'+'.join(f'IFERROR({l}{row},0)' for l in rl)}))"
        else:
            formula=f"=MIN(10,MAX(0,{'+'.join(f'IFERROR({l}{row},0)' for l in rl)}-IF({pr}{row}=1,2,0)))"

        rb_base=s["auto_base"]
        rbg_rel="green" if rb_base>=7 else ("amber" if rb_base>=4 else rbg)
        c=ws.cell(row,CI_TOT,formula)
        c.fill=fp(rbg_rel); c.font=df_(bold=True); c.border=bdr(); c.alignment=ac()

        # Pregnant
        pf=s["preg_flag"]
        if pf:
            c=wc(CI_PREG,"⚠ Pregnant",bg="red",bold=True)
            c.font=Font(name=FN,bold=True,color="C00000",size=8)
        else:
            wc(CI_PREG,0,bg=rbg)

        # Notes
        note_val = m.get("notes2") or m.get("award_name") or s.get("kids_age","")
        note_mm  = "; ".join(s.get("mismatches",[])) if s.get("mismatches") else ""
        manual_cell(CI_N1, m.get("notes1"))
        wc(CI_N2, note_val or note_mm, bg=rbg, left=True)

    # Legend
    lg=ROW_D+len(scores)+2
    ws.cell(lg,1,"LEGEND").font=Font(name=FN,bold=True,size=9)
    for i,(clr,txt) in enumerate([
        ("manual","⬛ Soft Yellow = Manual field | Dark green text = pre-filled from Grading sheet"),
        ("green", "🟢 Green  = Pass / Clean (auto-scored)"),
        ("amber", "🟠 Amber  = Review (auto-scored)"),
        ("red",   "🔴 Pink   = Flagged (auto-scored)"),
        ("brtred","🔴 Bright Red = Auto-reject or Risk Score ≥ 5"),
    ]):
        ws.cell(lg+1+i,1,"").fill=fp(clr); ws.cell(lg+1+i,1,"").border=bdr()
        ws.cell(lg+1+i,2,txt).font=Font(name=FN,size=8)
    ws.freeze_panes=f"C{ROW_D}"

# ─────────────────────────────────────────────────────────────────────────────
# COMPOSITE GRADING SHEET
# Shows Risk MAX + Relevance AVG per influencer across all platforms
# ─────────────────────────────────────────────────────────────────────────────

def write_composite_sheet(ws, all_platform_scores, use_case, profiles):
    ws.row_dimensions[1].height = 22
    ws.cell(1,1,f"Composite Grading — All Platforms ({use_case})").font = \
        Font(name=FN,bold=True,size=12,color=BG["navy"])

    COLS = [
        ("Influencer Name",                "navy",  22),
        ("Handle",                         "navy",  18),
        ("Risk MAX\n(all platforms)",       "dkred", 16),
        ("Relevance AVG\n(auto-base)",      "dkgrn", 18),
        ("Risk Zone",                      "navy",  16),
        ("Relevance Zone",                 "navy",  16),
        ("Recommendation",                 "navy",  20),
        ("Platforms",                      "navy",  28),
        ("Total Videos\n(all platforms)",  "navy",  16),
    ]
    for i,(_,_,w) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ROW_H,ROW_D=3,4; ws.row_dimensions[ROW_H].height=42
    for ci,(lbl,bg,_) in enumerate(COLS,1):
        c=ws.cell(ROW_H,ci,lbl); c.fill=fp(bg); c.font=hf(); c.border=bdr(); c.alignment=ac()

    # Aggregate
    by_h = {}
    for plat,sl in all_platform_scores.items():
        for s in sl:
            h=s["handle"]
            if h not in by_h: by_h[h]={"risks":[],"rels":[],"plats":[],"vids":0}
            by_h[h]["risks"].append(s["risk_score"])
            by_h[h]["rels"].append(s["auto_base"])
            by_h[h]["plats"].append(plat)
            by_h[h]["vids"] += s.get("total_videos",0)

    for off,(handle,d) in enumerate(sorted(by_h.items())):
        row=ROW_D+off; rbg="white" if off%2==0 else "grey"
        ws.row_dimensions[row].height=16
        mr=max(d["risks"]); al2=round(sum(d["rels"])/len(d["rels"]),2)
        plats=", ".join(sorted(set(d["plats"])))
        r_bg="red" if mr>=5 else ("amber" if mr>0 else "green")
        l_bg="green" if al2>=8 else ("amber" if al2>=4 else "red")
        rec="❌ Reject" if mr>=5 else ("✅ Approve" if mr==0 and al2>=7 else "🟡 Manual Review")
        rc="red" if "Reject" in rec else ("green" if "Approve" in rec else "amber")

        def mc(col,val,bg,bold=False,left=False):
            c=ws.cell(row,col,val); c.fill=fp(bg); c.font=df_(bold=bold)
            c.border=bdr(); c.alignment=al() if left else ac()

        mc(1,profiles.get(handle,handle),"name",bold=True,left=True)
        mc(2,handle,"name",left=True)
        mc(3,mr,r_bg,bold=True); mc(4,al2,l_bg,bold=True)
        mc(5,"🔴 Reject" if mr>=5 else "🟠 Review" if mr>0 else "🟢 Clean",r_bg)
        mc(6,"🟢 High" if al2>=8 else "🟠 Moderate" if al2>=4 else "🔴 Low",l_bg)
        mc(7,rec,rc,bold=True); mc(8,plats,rbg)
        mc(9,d["vids"],rbg)

    nr=ROW_D+len(by_h)+2
    ws.cell(nr,1,"NOTES").font=Font(name=FN,bold=True,size=9)
    for i,n in enumerate([
        "Risk = MAX across all platforms | Relevance = AVERAGE across all platforms (auto-base only).",
        "Relevance auto-base excludes manual fields. Fill yellow cells in platform sheets for final total.",
        "⬛ Yellow / Dark-green cells in platform grading sheets = manual input (dark green = pre-filled).",
    ]): ws.cell(nr+1+i,1,n).font=Font(name=FN,size=8,italic=True,color="595959")
    ws.freeze_panes=f"C{ROW_D}"

# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

PLAT_LABELS={"TT":"TikTok","FB":"Facebook","YT":"YouTube","IG_YT":"IG + YT","IG":"Instagram","Captions":"Captions"}

def build_excel(platform_scores_map, use_case, profiles):
    """
    platform_scores_map: dict label → (scores_list, manual_scores_dict)
      label can be "TT", "FB", "IG", "Captions File", or any user-defined name
    """
    wb=openpyxl.Workbook()

    # Composite sheet first
    ws_comp=wb.active; ws_comp.title="Composite Grading Sheet"
    all_for_composite={lbl: sc for lbl,(sc,_) in platform_scores_map.items()}
    write_composite_sheet(ws_comp, all_for_composite, use_case, profiles)

    # Per-platform grading sheets
    for label,(scores,manual) in platform_scores_map.items():
        if not scores: continue
        safe_title = f"{label} Grading Sheet"[:31]
        ws=wb.create_sheet(title=safe_title)
        write_grading_sheet(ws, scores, label, use_case, profiles, manual_scores=manual)

    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf


# ─────────────────────────────────────────────────────────────────────────────
# INFLUENCER RAW SHEET WRITER  (with formula rows)
# ─────────────────────────────────────────────────────────────────────────────

# Code/Point formula template per column name.
# {L} = placeholder for the actual Excel column letter.
# Kids/Topics/Media: use row 2 (count) directly since binary presence check.
# Risk params: use row 4 (% prevalence) with threshold formulas.
FORMULA_CODE_POINT = {
    "Kids Presenence":                   '=IF({L}2>0,"GREEN","RED")',
    "Brand Partnership Presence":         "Check Manually for Competition",
    "Media Presence & Awards":           '=IF({L}2>0,"GREEN","RED")',
    "Alcohol Use Discussion":            '=IF({L}4>0.15,"RED",IF({L}4>0.05,"AMBER","GREEN"))',
    "Vaccination":                       '=IF({L}4>0,"RED","GREEN")',
    "Health care stance":                '=IF({L}4>0,"RED","GREEN")',
    "Ultra-processed food":              '=IF({L}4>0.3,"RED",IF({L}4>0.15,"AMBER","GREEN"))',
    "Pregnant":                          '=IF({L}4>0,"RED","GREEN")',
    "Profanity Usage":                   '=IF({L}4>0.3,"RED",IF({L}4>0.15,"AMBER","GREEN"))',
    "Substance Use Discussion":          '=IF({L}4>0,"RED","GREEN")',
    "Sensitive Visual Content":          '=IF({L}4>0.2,"RED","GREEN")',
    "Stereotypes or Bias":               '=IF({L}4>0.15,"RED",IF({L}4>0.05,"AMBER","GREEN"))',
    "Violence Advocacy":                 '=IF({L}4>=0.05,"RED",IF({L}4>0,"AMBER","GREEN"))',
    "Political Stance":                  '=IF({L}4>0.25,"RED",IF({L}4>0.1,"AMBER","GREEN"))',
    "Breastfeeding":                     '=IF({L}4>0,"RED","GREEN")',
    "Unscientific Claims":               '=IF({L}4>=0.15,"RED","GREEN")',
    "Topics around Adult Health":        '=IF({L}2>0,"GREEN","RED")',
    "Topics on Adult Healthy Nutrition": '=IF({L}2>0,"GREEN","RED")',
}

# Background colours for the 4 formula rows
ROW_COLORS = {
    2: "D9E1F2",   # light blue  — YES count
    3: "EBF3FB",   # lighter blue — total videos
    4: "FFF2CC",   # light yellow — % prevalence
    5: "E2EFDA",   # light green  — Code/Point
}


def write_influencer_sheet(wb, df, sheet_title, fmt):
    """
    Write one influencer's raw data sheet into the workbook.

    Structure written (matches the reference Excel exactly):
      Row 1 : column headers
      Row 2 : HEADER label  +  COUNTIFS("Yes") formula per flag column
      Row 3 : Total Videos Processed  +  COUNTA formula per flag column
      Row 4 : % Prevalence  +  =row2/row3 per flag column
      Row 5 : Code/Point  +  IF threshold formula per flag column
      Row 6+: individual video rows (Yes/No data)

    For Format A DataFrames, the summary rows are already in pandas rows 0-3
    and data starts at pandas row 4. We skip those and rewrite from scratch so
    formulas are always correct (fixing any typos from the source file).

    For Format B DataFrames (Yes/No individual rows), data starts at pandas
    row 0 and there are no existing summary rows — we add them here.
    """
    safe_title = sheet_title[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=safe_title)

    # ── Determine data rows ───────────────────────────────────────────────
    # Format A has 4 summary rows at the top (pandas rows 0-3), data from row 4.
    # Format B has data from row 0.
    data_start_pandas = 4 if fmt in ("A", "C", "IG") else 0
    data_rows = df.iloc[data_start_pandas:].reset_index(drop=True)

    col_names = list(df.columns)
    n_cols    = len(col_names)

    # Build col_name → Excel column letter mapping
    col_letter = {}
    for i, name in enumerate(col_names):
        col_letter[name] = get_column_letter(i + 1)

    # Excel data starts at row 6 (rows 1-5 are header + 4 formula rows)
    data_excel_start = 6

    # ── Row 1: column headers ──────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name=FN, bold=True, color="FFFFFF", size=9)
    ws.row_dimensions[1].height = 20
    for i, name in enumerate(col_names, 1):
        c = ws.cell(1, i, name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(str(name)) + 2))

    # ── Rows 2–5: formula rows ─────────────────────────────────────────────
    row_labels = {
        2: ("response_1", "HEADER"),
        3: ("response_1", "Total Videos Processed"),
        4: ("response_1", "% Prevalence"),
        5: ("response_1", "Code/Point"),
    }
    label_fill = PatternFill("solid", fgColor="2E75B6")
    label_font = Font(name=FN, bold=True, color="FFFFFF", size=9)

    for row_num in [2, 3, 4, 5]:
        ws.row_dimensions[row_num].height = 16
        row_bg = PatternFill("solid", fgColor=ROW_COLORS[row_num])
        row_font = Font(name=FN, bold=True, size=9, color="000000")

        # Label in response_1 col (col B typically)
        label_col, label_text = row_labels[row_num]
        if label_col in col_letter:
            lc = ws.cell(row_num, col_names.index(label_col) + 1, label_text)
            lc.fill = label_fill; lc.font = label_font
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.border = bdr()

        for col_name, letter in col_letter.items():
            col_idx = col_names.index(col_name) + 1
            cell = ws.cell(row_num, col_idx)
            cell.fill = row_bg
            cell.font = row_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bdr()

            if col_name not in FORMULA_CODE_POINT and col_name != "response_1":
                # No formula for this column — leave empty
                continue

            if row_num == 2 and col_name in FORMULA_CODE_POINT:
                # YES count
                cell.value = f'=COUNTIFS({letter}{data_excel_start}:{letter}10000,"Yes")'
                cell.number_format = "0"

            elif row_num == 3 and col_name in FORMULA_CODE_POINT:
                # Total videos — use COUNTA of same column
                cell.value = f'=COUNTA({letter}{data_excel_start}:{letter}10000)'
                cell.number_format = "0"

            elif row_num == 4 and col_name in FORMULA_CODE_POINT:
                # % prevalence
                cell.value = f'=IFERROR({letter}2/{letter}3,0)'
                cell.number_format = "0.00%"

            elif row_num == 5 and col_name in FORMULA_CODE_POINT:
                # Code/Point — apply threshold formula
                template = FORMULA_CODE_POINT[col_name]
                if template.startswith("="):
                    cell.value = template.replace("{L}", letter)
                else:
                    cell.value = template  # e.g. "Check Manually for Competition"
                # Colour the cell based on what it will evaluate to
                # (we can't know at write-time so use neutral)
                cell.fill = PatternFill("solid", fgColor="E2EFDA")

    # ── Rows 6+: data rows (use append for speed) ─────────────────────────
    alt_fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F2F2F2"),
    ]
    data_font  = Font(name=FN, size=8)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
    data_border= bdr()

    for row_offset, (_, data_row) in enumerate(data_rows.iterrows()):
        excel_row = data_excel_start + row_offset
        row_fill  = alt_fills[row_offset % 2]
        ws.row_dimensions[excel_row].height = 14

        # Build list of values for this row
        row_vals = []
        for col_name in col_names:
            val = data_row[col_name]
            row_vals.append(None if pd.isna(val) else val)

        # Append the row values (fast)
        ws.append(row_vals)

        # Apply minimal styling to the row (skip border on bulk rows for speed)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(excel_row, col_idx)
            cell.fill = row_fill
            cell.font = data_font

    # Freeze header + formula rows
    ws.freeze_panes = f"A{data_excel_start}"
    return ws


def build_excel_with_sheets(platform_scores_map, use_case, profiles,
                             raw_xl_map=None):
    """
    Build the complete output Excel file.

    platform_scores_map : dict  label → (scores_list, manual_dict)
    raw_xl_map          : dict  label → xl_dict  (raw Excel data per platform)
                          Optional. If provided, each influencer's raw sheet is
                          written with formula rows into the output file.
    """
    wb = openpyxl.Workbook()

    # ── 1. Composite Grading Sheet ────────────────────────────────────────
    ws_comp = wb.active
    ws_comp.title = "Composite Grading Sheet"
    all_for_composite = {lbl: sc for lbl, (sc, _) in platform_scores_map.items()}
    write_composite_sheet(ws_comp, all_for_composite, use_case, profiles)

    # ── 2. Per-platform Grading Sheets ────────────────────────────────────
    for label, (scores, manual) in platform_scores_map.items():
        if not scores: continue
        ws = wb.create_sheet(title=f"{label} Grading Sheet"[:31])
        write_grading_sheet(ws, scores, label, use_case, profiles,
                            manual_scores=manual)

    # ── 3. Individual influencer sheets with formula rows ─────────────────
    if raw_xl_map:
        for platform_label, xl in raw_xl_map.items():
            # Detect which sheets belong to this platform
            is_cap = "Caption" in platform_label
            if is_cap:
                sheet_list = detect_captions_sheets(xl)
            else:
                # Map UI label → internal key
                plat_key_map = {
                    "TikTok": "TT", "Facebook": "FB",
                    "YouTube": "YT", "YouTube (YT)": "YT",
                    "Instagram": "IG", "IG + YouTube": "IG_YT",
                }
                pk = plat_key_map.get(platform_label, platform_label)
                grp = detect_sheets(xl)
                sheet_list = grp.get(pk, [])

            for sn, handle in sheet_list:
                df  = xl[sn]
                fmt = detect_format(df)
                # Sheet title: use original sheet name, truncated to 31 chars
                sheet_title = sn[:31]
                try:
                    write_influencer_sheet(wb, df, sheet_title, fmt)
                except Exception as e:
                    # Never crash the whole build for one bad sheet
                    pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Abbott Influencer Vetting",page_icon="🏥",layout="wide")
st.markdown("""<style>
.stApp{font-family:Arial,sans-serif;}h1{color:#1F3864;}h2{color:#2E75B6;}
.g{color:#375623;font-weight:bold;}.a{color:#806000;font-weight:bold;}.r{color:#C00000;font-weight:bold;}
div[data-testid='stExpander']{border:1px solid #e0e0e0;border-radius:6px;}
</style>""",unsafe_allow_html=True)

st.title("🏥 Abbott Influencer Vetting — Auto-Grader v6.0")
st.caption("Upload multiple platform files → auto-score → composite grading | Pre-filled manual scores supported")
st.divider()

# ── Dynamic multi-file upload ──────────────────────────────────────────────
st.subheader("📂 Upload Platform Files")
st.caption("Add one file per platform. Label each file with its platform. Main AI files + optional Captions file.")

PLATFORM_OPTIONS = ["TikTok (TT)", "Facebook (FB)", "YouTube (YT)", "Instagram (IG)", "IG + YouTube",
                    "Captions File (handle-named sheets)", "Other / Mixed"]

# Session state for file slots
if "num_slots" not in st.session_state:
    st.session_state.num_slots = 1

col_add, col_reset = st.columns([1,4])
with col_add:
    if st.button("➕ Add another file"):
        st.session_state.num_slots = min(st.session_state.num_slots + 1, 8)
with col_reset:
    if st.button("🗑 Reset all"):
        st.session_state.num_slots = 1

uploaded_files = []   # list of (platform_label, file_obj)
for i in range(st.session_state.num_slots):
    c1, c2 = st.columns([1, 3])
    with c1:
        plat = st.selectbox(f"Platform {i+1}", PLATFORM_OPTIONS, key=f"plat_{i}")
    with c2:
        f = st.file_uploader(f"File {i+1}", type=["xlsx"], key=f"file_{i}",
                             label_visibility="collapsed")
    if f:
        uploaded_files.append((plat, f))

st.divider()

if not uploaded_files:
    st.info("👆 Upload at least one platform file to begin.")
    with st.expander("ℹ️ How it works"):
        st.markdown("""
**Upload multiple platform files:**
Each file can contain influencer sheets for one platform (TT, FB, IG etc.) or mixed.
Label each with its platform — this becomes the tab name in the output Excel.

**Auto-detects:**
- Use case (Pediatric vs Adult) from column content
- Sheet format (aggregated vs Yes/No rows vs FB Verdict format)
- Pre-filled manual scores from any existing Grading sheet in the file

**Scoring rules:**
- Auto-reject gates (Substance/BF/Vax/HC/Competition): value = **10** if triggered, 0 if clean
- 8 risk params: % threshold → GREEN/AMBER/RED → 0/0.5/1.0 pts → Risk Score = (sum/8)×10
- Relevance: count > 0 → GREEN → full pts | count = 0 → RED → 0 pts

**⬛ Yellow cells = manual input | 🟩 Dark green text = pre-filled from your Grading sheet**

**Output sheets:** Composite Grading Sheet + one grading tab per uploaded platform file
        """)

else:
    with st.spinner("Reading files…"):
        # Determine use case from first file that can detect it
        use_case = "Pediatric"
        profiles = {}
        for _, f in uploaded_files:
            xl_tmp = pd.read_excel(f, sheet_name=None)
            f.seek(0)
            uc_tmp = detect_use_case(xl_tmp)
            if uc_tmp == "Adult": use_case = "Adult"; break
            prf_tmp = build_profiles(xl_tmp)
            if prf_tmp: profiles.update(prf_tmp)

    st.success(f"✅ {len(uploaded_files)} file(s) loaded | Use case: **{use_case}**")
    st.divider()

    # ── Score each file ────────────────────────────────────────────────────
    platform_scores_map = {}   # label → (scores_list, manual_dict)
    raw_xl_map          = {}   # label → xl dict (for influencer sheets)
    all_warns = []

    for plat_label, f in uploaded_files:
        f.seek(0)
        xl = pd.read_excel(f, sheet_name=None)
        prf = build_profiles(xl)
        if prf: profiles.update(prf)

        # Store raw xl for influencer sheet writing
        raw_xl_map[plat_label.split("(")[0].strip()] = xl

        # Read manual scores if Grading sheet exists
        manual = read_manual_scores(xl)

        is_cap_file = "Captions File" in plat_label

        with st.spinner(f"Scoring {plat_label}…"):
            if is_cap_file:
                scores, warns = score_file(xl, plat_label, use_case, is_captions_file=True)
            else:
                plat_key_map = {
                    "TikTok (TT)":    "TT",
                    "Facebook (FB)":  "FB",
                    "YouTube (YT)":   "YT",
                    "Instagram (IG)": "IG",
                    "IG + YouTube":   "IG_YT",
                    "Other / Mixed":  None,
                }
                pk = plat_key_map.get(plat_label)
                if pk:
                    scores, warns = score_file(xl, pk, use_case, is_captions_file=False)
                else:
                    grp = detect_sheets(xl)
                    scores, warns = [], []
                    for p, sl in grp.items():
                        for sn, h in sl:
                            df = xl[sn]
                            try:
                                s = score_sheet(df, h, use_case)
                                actual = extract_handle(df)
                                if actual: s["handle"] = actual
                                scores.append(s)
                            except Exception as e:
                                warns.append(f"{sn}: {e}")

        all_warns.extend(warns)
        display = plat_label.split("(")[0].strip()
        platform_scores_map[display] = (scores, manual)
        st.write(f"  ✅ **{display}**: {len(scores)} sheets scored"
                 + (f" | {len(manual)} pre-filled manual records found" if manual else "")
                 + (f" | ⚠ {len(warns)} errors" if warns else ""))

    if all_warns:
        with st.expander(f"⚠ {len(all_warns)} warnings"):
            for w in all_warns: st.text(w)

    # ── Preview composite ──────────────────────────────────────────────────
    st.divider()
    st.subheader("📊 Composite Preview")

    rows = []
    by_h = {}
    for lbl,(sl,_) in platform_scores_map.items():
        for s in sl:
            h=s["handle"]
            if h not in by_h: by_h[h]={"risks":[],"rels":[],"plats":[]}
            by_h[h]["risks"].append(s["risk_score"])
            by_h[h]["rels"].append(s["auto_base"])
            by_h[h]["plats"].append(lbl)

    for h,d in sorted(by_h.items()):
        mr=max(d["risks"]); al2=round(sum(d["rels"])/len(d["rels"]),2)
        rec="❌ Reject" if mr>=5 else ("✅ Approve" if mr==0 and al2>=7 else "🟡 Review")
        rows.append({"Name":profiles.get(h,h),"Handle":h,"Risk MAX":mr,
                     "Rel AVG (auto)":al2,
                     "Platforms":", ".join(sorted(set(d["plats"]))),"Rec":rec})

    dfp=pd.DataFrame(rows)
    if not dfp.empty:
        def sty(row):
            s=[""]*len(row)
            ri=dfp.columns.get_loc("Risk MAX"); li=dfp.columns.get_loc("Rel AVG (auto)")
            rci=dfp.columns.get_loc("Rec")
            s[ri]=f"background-color:{'#FFC7CE' if row['Risk MAX']>=5 else '#FFEB9C' if row['Risk MAX']>0 else '#C6EFCE'}"
            s[li]=f"background-color:{'#C6EFCE' if row['Rel AVG (auto)']>=7 else '#FFEB9C' if row['Rel AVG (auto)']>=4 else '#FFC7CE'}"
            s[rci]=f"background-color:{'#C6EFCE' if 'Approve' in row['Rec'] else '#FFC7CE' if 'Reject' in row['Rec'] else '#FFEB9C'};font-weight:bold"
            return s
        st.dataframe(dfp.style.apply(sty,axis=1),use_container_width=True,
                     height=min(500,50+len(rows)*36))

        appr=sum(1 for r in rows if "Approve" in r["Rec"])
        rev=sum(1 for r in rows if "Review" in r["Rec"])
        rej=sum(1 for r in rows if "Reject" in r["Rec"])
        st.markdown(
            f"**{len(rows)} influencers across {len(platform_scores_map)} platform(s)** &nbsp;|&nbsp; "
            f"<span class='g'>✅ Approve: {appr}</span> &nbsp;|&nbsp; "
            f"<span class='a'>🟡 Review: {rev}</span> &nbsp;|&nbsp; "
            f"<span class='r'>❌ Reject: {rej}</span>",
            unsafe_allow_html=True
        )

    # ── Download ───────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📥 Download Graded Excel")
    st.caption("Contains: Composite Grading + per-platform Grading Sheets + every influencer's raw sheet with formula rows")

    with st.spinner("Building output Excel (writing formula rows for all influencer sheets)…"):
        buf = build_excel_with_sheets(platform_scores_map, use_case, profiles,
                                      raw_xl_map=raw_xl_map)

    first_fname = uploaded_files[0][1].name.replace(".xlsx","") if uploaded_files else "Abbott"
    fname = f"{first_fname}_GRADED_{use_case}.xlsx"

    st.download_button(
        "⬇️ Download Complete Graded Excel",
        data=buf, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )

    sheet_list = ["Composite Grading Sheet"] + [f"{lbl} Grading Sheet" for lbl in platform_scores_map if platform_scores_map[lbl][0]]
    total_inf_sheets = sum(len(sc) for sc, _ in platform_scores_map.values())
    st.caption(
        f"**Output:** {' | '.join(sheet_list)} | "
        f"+ {total_inf_sheets} influencer raw sheets (each with 4 formula rows: Count / Total / % / Code-Point)  |  "
        "⬛ Yellow = manual input | 🟩 Dark green = pre-filled"
    )
